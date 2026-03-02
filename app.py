import os
from datetime import datetime
import random
import string
import re
import io
import secrets
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bson.objectid import ObjectId

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, session, jsonify, make_response
)
from flask_pymongo import PyMongo
from flask_login import (
    LoginManager, UserMixin, login_user,
    login_required, logout_user, current_user
)
from flask_mail import Mail, Message
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from itsdangerous import URLSafeTimedSerializer
from config import Config

app = Flask(__name__)
app.config.from_object(Config)

app.wsgi_app = ProxyFix(
    app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1
)

s = URLSafeTimedSerializer(app.config['SECRET_KEY'])

mongo = PyMongo(app)
mail = Mail(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = "Sessão expirada. Por favor, autentique-se novamente."
login_manager.login_message_category = "error"


def validate_cpf(cpf):
    cpf = re.sub(r'[^0-9]', '', cpf)
    if len(cpf) != 11: return False
    if cpf == cpf[0] * 11: return False
    soma = sum(int(cpf[i]) * (10 - i) for i in range(9))
    resto = (soma * 10) % 11
    if resto == 10 or resto == 11: resto = 0
    if resto != int(cpf[9]): return False
    soma = sum(int(cpf[i]) * (11 - i) for i in range(10))
    resto = (soma * 10) % 11
    if resto == 10 or resto == 11: resto = 0
    if resto != int(cpf[10]): return False
    return True


def validate_cnpj(cnpj):
    cnpj = re.sub(r'[^0-9]', '', cnpj)
    if len(cnpj) != 14: return False
    if cnpj == cnpj[0] * 14: return False
    tamanho = 12
    numeros = cnpj[:tamanho]
    digitos = cnpj[tamanho:]
    soma = 0
    pos = tamanho - 7
    for i in range(tamanho, 0, -1):
        soma += int(numeros[tamanho - i]) * pos
        pos -= 1
        if pos < 2: pos = 9
    resultado = 0 if soma % 11 < 2 else 11 - (soma % 11)
    if resultado != int(digitos[0]): return False
    tamanho = 13
    numeros = cnpj[:tamanho]
    soma = 0
    pos = tamanho - 7
    for i in range(tamanho, 0, -1):
        soma += int(numeros[tamanho - i]) * pos
        pos -= 1
        if pos < 2: pos = 9
    resultado = 0 if soma % 11 < 2 else 11 - (soma % 11)
    if resultado != int(digitos[1]): return False
    return True


def get_email_template(preheader, title, content_html, button_text=None, button_link=None):
    html = f"""
    <!DOCTYPE html>
    <html lang="pt-br">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            body {{ margin: 0; padding: 0; font-family: 'Helvetica Neue', Helvetica, Arial, sans-serif; background-color: #F3F4F6; color: #413D3A; }}
            .email-wrapper {{ width: 100%; background-color: #F3F4F6; padding: 40px 0; }}
            .email-container {{ max-width: 600px; margin: 0 auto; background-color: #ffffff; border-radius: 16px; overflow: hidden; box-shadow: 0 10px 30px rgba(0,0,0,0.08); }}
            .header {{ background-color: #ffffff; padding: 35px 40px; text-align: center; border-bottom: 5px solid #FBBA00; }}
            .logo-img {{ max-height: 50px; width: auto; display: block; margin: 0 auto; }}
            .content {{ padding: 45px 40px; background-color: #ffffff; }}
            .email-title {{ color: #413D3A; font-size: 24px; font-weight: 800; margin-top: 0; margin-bottom: 25px; letter-spacing: -0.5px; line-height: 1.3; }}
            .email-text {{ font-size: 16px; line-height: 1.6; color: #555555; margin-bottom: 25px; }}
            .btn-container {{ text-align: center; margin: 40px 0 20px 0; }}
            .btn {{ background-color: #FBBA00; color: #413D3A; padding: 18px 45px; text-decoration: none; font-weight: 800; border-radius: 10px; font-size: 14px; text-transform: uppercase; letter-spacing: 1px; display: inline-block; box-shadow: 0 4px 15px rgba(251, 186, 0, 0.4); }}
            .footer {{ background-color: #f9fafb; padding: 30px 40px; text-align: center; border-top: 1px solid #eeeeee; font-size: 12px; color: #999; }}
            .highlight-box {{ background-color: #fffbf0; border: 1px solid #ffeeba; border-left: 4px solid #FBBA00; padding: 25px; border-radius: 8px; margin: 30px 0; text-align: center; }}
            .token-code {{ font-family: 'Courier New', monospace; font-size: 36px; font-weight: 700; color: #413D3A; letter-spacing: 8px; display: block; margin: 0; }}
            .receipt-table {{ width: 100%; margin: 25px 0; border: 1px solid #eeeeee; border-radius: 8px; overflow: hidden; border-collapse: collapse; }}
            .receipt-row td {{ padding: 15px; border-bottom: 1px solid #eeeeee; font-size: 14px; color: #413D3A; }}
            .receipt-row:nth-child(even) {{ background-color: #fcfcfc; }}
            .receipt-value {{ font-weight: 700; text-align: right; white-space: nowrap; }}
        </style>
    </head>
    <body>
        <div class="email-wrapper">
            <div style="display:none;font-size:1px;color:#333;line-height:1px;max-height:0px;max-width:0px;opacity:0;overflow:hidden;">{preheader}</div>
            <div class="email-container">
                <div class="header">
                    <img src="cid:logo" alt="Scryta" class="logo-img">
                </div>
                <div class="content">
                    <h1 class="email-title">{title}</h1>
                    <div class="email-text">{content_html}</div>
                    {f'<div class="btn-container"><a href="{button_link}" class="btn">{button_text}</a></div>' if button_link else ''}
                </div>
                <div class="footer">
                    <p>&copy; {datetime.now().year} Scryta Contabilidade Digital.</p>
                </div>
            </div>
        </div>
    </body>
    </html>
    """
    return html


def send_email_with_logo(msg):
    try:
        with app.open_resource("static/img/logo.png") as fp:
            msg.attach("logo.png", "image/png", fp.read(),
                       headers={'Content-ID': '<logo>', 'Content-Disposition': 'inline'})
    except Exception:
        pass
    mail.send(msg)


def log_action(user_id, user_email, action, details=None):
    mongo.db.logs.insert_one({
        "user_id": ObjectId(user_id) if user_id else None,
        "email": user_email,
        "action": action,
        "details": details,
        "ip": request.remote_addr,
        "timestamp": datetime.utcnow()
    })


class User(UserMixin):
    def __init__(self, user_data):
        self.id = str(user_data['_id'])
        self.email = user_data['email']
        self.name = user_data['name']
        self.cpf = user_data.get('cpf')
        self.is_admin = user_data.get('is_admin', False)
        self.term_accepted_at = user_data.get('term_accepted_at')


@login_manager.user_loader
def load_user(user_id):
    try:
        user_data = mongo.db.users.find_one({"_id": ObjectId(user_id)})
        return User(user_data) if user_data else None
    except:
        return None


def is_password_strong(password):
    if len(password) < 8: return False
    if not re.search(r"[A-Z]", password): return False
    if not re.search(r"\d", password): return False
    return True


def generate_token():
    return ''.join(secrets.choice(string.digits) for _ in range(6))


@app.route('/')
def index():
    if current_user.is_authenticated: return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        email = request.form.get('email')
        if mongo.db.users.find_one({"email": email}):
            flash('Este e-mail corporativo já consta em nossa base.', 'error')
            return redirect(url_for('login'))

        token = s.dumps(email, salt='email-confirm')
        link = url_for('register_complete', token=token, _external=True)

        try:
            content = "<p>Recebemos uma solicitação para integrar sua empresa ao ecossistema Scryta.</p>"
            msg = Message('Convite Empresarial - Scryta', recipients=[email])
            msg.html = get_email_template("Ative sua conta.", "Bem-vindo", content, "ATIVAR CONTA", link)
            send_email_with_logo(msg)
            flash(f'Link enviado para {email}.', 'success')
        except Exception as e:
            flash(f'Erro SMTP: {str(e)}', 'error')

        return redirect(url_for('login'))
    return render_template('signup.html')


@app.route('/register/<token>', methods=['GET', 'POST'])
def register_complete(token):
    try:
        email = s.loads(token, salt='email-confirm', max_age=3600)
    except:
        flash('Link inválido ou expirado.', 'error')
        return redirect(url_for('signup'))

    if request.method == 'POST':
        name = request.form.get('name')
        cpf = request.form.get('cpf')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        terms = request.form.get('terms')

        if not terms:
            flash('Aceite os termos.', 'error')
            return redirect(request.url)
        if not validate_cpf(cpf):
            flash('CPF inválido.', 'error')
            return redirect(request.url)
        if password != confirm_password:
            flash('Senhas não conferem.', 'error')
            return redirect(request.url)
        if not is_password_strong(password):
            flash('Senha fraca.', 'error')
            return redirect(request.url)
        if mongo.db.users.find_one({"cpf": cpf}):
            flash('CPF já cadastrado.', 'error')
            return redirect(request.url)

        hashed = generate_password_hash(password)
        is_admin = True if mongo.db.users.count_documents({}) == 0 else False
        user_id = mongo.db.users.insert_one({
            "name": name, "cpf": cpf, "email": email, "password": hashed,
            "is_admin": is_admin, "created_at": datetime.utcnow(),
            "term_accepted_at": datetime.utcnow(), "status": "active"
        }).inserted_id

        log_action(user_id, email, "REGISTER", f"CPF: {cpf}")
        flash('Cadastro concluído!', 'success')
        return redirect(url_for('login'))
    return render_template('register.html', email=email, token=token)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        user_data = mongo.db.users.find_one({"email": email})

        if user_data and check_password_hash(user_data['password'], password):
            user = User(user_data)
            login_user(user)
            session.pop('batch_items', None)
            return redirect(url_for('dashboard'))
        else:
            flash('Credenciais inválidas.', 'error')
    return render_template('login.html')


@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email')
        user = mongo.db.users.find_one({"email": email})
        if user:
            token = s.dumps(email, salt='password-reset')
            link = url_for('reset_password', token=token, _external=True)
            try:
                content = "<p>Você solicitou a redefinição de sua senha de acesso.</p>"
                msg = Message('Redefinição de Senha - Scryta', recipients=[email])
                msg.html = get_email_template("Redefinir Senha", "Recuperação", content, "CRIAR NOVA SENHA", link)
                send_email_with_logo(msg)
            except Exception as e:
                pass
        flash('Se o e-mail existir em nossa base, um link de recuperação foi enviado.', 'success')
        return redirect(url_for('login'))
    return render_template('forgot_password.html')


@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    try:
        email = s.loads(token, salt='password-reset', max_age=3600)
    except:
        flash('Link inválido ou expirado.', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if password != confirm_password:
            flash('Senhas não conferem.', 'error')
            return redirect(request.url)
        if not is_password_strong(password):
            flash('Senha não atende aos critérios de segurança.', 'error')
            return redirect(request.url)

        hashed = generate_password_hash(password)
        mongo.db.users.update_one({"email": email}, {"$set": {"password": hashed}})
        flash('Senha atualizada com sucesso!', 'success')
        return redirect(url_for('login'))
    return render_template('reset_password.html', token=token)


@app.route('/logout')
@login_required
def logout():
    session.clear()
    logout_user()
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.is_admin: return redirect(url_for('admin_panel'))
    
    user_companies = list(mongo.db.companies.find({"user_id": ObjectId(current_user.id)}))
    user_partners = list(mongo.db.partners.find({"user_id": ObjectId(current_user.id)}))
    user_history = list(mongo.db.user_financials.find({"user_id": ObjectId(current_user.id)}).sort("submitted_at", -1))
    
    response = make_response(render_template('dashboard.html', 
                                           name=current_user.name, 
                                           companies=user_companies,
                                           partners=user_partners,
                                           history=user_history))
    
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    
    return response


@app.route('/add_company', methods=['POST'])
@login_required
def add_company():
    company_name = request.form.get('company_name') 
    company_cnpj = request.form.get('company_cnpj')

    if not company_cnpj:
        flash('CNPJ obrigatório.', 'error')
        return redirect(url_for('dashboard'))
    if not validate_cnpj(company_cnpj):
        flash('CNPJ inválido.', 'error')
        return redirect(url_for('dashboard'))

    exists = mongo.db.companies.find_one({"user_id": ObjectId(current_user.id), "cnpj": company_cnpj})
    if exists:
        flash('Empresa já cadastrada.', 'error')
    else:
        final_name = company_name if company_name else company_cnpj
        mongo.db.companies.insert_one({
            "user_id": ObjectId(current_user.id), "name": final_name,
            "cnpj": company_cnpj, "created_at": datetime.utcnow()
        })
        flash('Empresa adicionada!', 'success')
    return redirect(url_for('dashboard'))


@app.route('/add_partner', methods=['POST'])
@login_required
def add_partner():
    partner_name = request.form.get('partner_name')
    partner_cpf = request.form.get('partner_cpf')
    
    clean_cpf = re.sub(r'[^0-9]', '', partner_cpf)
    
    company_ids = request.form.getlist('company_ids')

    if not partner_name or not partner_cpf or not company_ids:
        flash('Preencha todos os campos do sócio e selecione ao menos uma empresa.', 'error')
        return redirect(url_for('dashboard'))
    
    if not validate_cpf(partner_cpf):
        flash('CPF do sócio inválido.', 'error')
        return redirect(url_for('dashboard'))

    added_count = 0
    for comp_id in company_ids:
        exists = mongo.db.partners.find_one({
            "user_id": ObjectId(current_user.id), 
            "company_id": comp_id, 
            "cpf": clean_cpf
        })
        
        if not exists:
            mongo.db.partners.insert_one({
                "user_id": ObjectId(current_user.id),
                "company_id": comp_id,
                "name": partner_name,
                "cpf": clean_cpf,
                "created_at": datetime.utcnow()
            })
            added_count += 1
            
    if added_count > 0:
        flash(f'Sócio adicionado a {added_count} empresa(s) com sucesso!', 'success')
    else:
        flash('O sócio já estava cadastrado na(s) empresa(s) selecionada(s).', 'error')
        
    return redirect(url_for('dashboard'))


@app.route('/request_cancel_token', methods=['POST'])
@login_required
def request_cancel_token():
    data = request.json
    record_id = data.get('record_id')
    
    record = mongo.db.user_financials.find_one({"_id": ObjectId(record_id), "user_id": ObjectId(current_user.id)})
    if not record:
        return jsonify({'status': 'error', 'message': 'Registro não encontrado.'}), 404

    token = generate_token()
    session['cancel_token'] = token
    session['cancel_record_id'] = record_id

    try:
        content = f"""
        <div class="highlight-box">
            <span style="font-size:12px; font-weight:bold; color:#999; text-transform:uppercase;">Código de Cancelamento</span>
            <span class="token-code">{token}</span>
        </div>
        <p style="text-align:center;">Este código valida o <strong>CANCELAMENTO</strong> do lançamento referente a <strong>{record.get('company_cnpj')}</strong> no valor de <strong>{record.get('valor')}</strong>.</p>
        """
        msg = Message("Token de Cancelamento - Scryta", recipients=[current_user.email])
        msg.html = get_email_template("Aviso de Segurança", "Cancelamento de Lançamento", content)
        send_email_with_logo(msg)
        return jsonify({'status': 'token_sent'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/record/cancel', methods=['POST'])
@login_required
def client_cancel_record():
    data = request.json
    code = data.get('code')
    record_id = session.get('cancel_record_id')
    
    if not code or not record_id or code != session.get('cancel_token'):
        return jsonify({'status': 'error', 'message': 'Token inválido ou expirado.'}), 400

    record = mongo.db.user_financials.find_one({"_id": ObjectId(record_id), "user_id": ObjectId(current_user.id)})
    if record:
        alerta = True if record.get('visualizado', False) else False
        mongo.db.user_financials.update_one(
            {"_id": ObjectId(record_id)},
            {"$set": {
                "status": "desconsiderado",
                "visualizado": False,
                "alerta_cancelamento": alerta,
                "cancelled_at": datetime.utcnow()
            }}
        )
        
        session.pop('cancel_token', None)
        session.pop('cancel_record_id', None)
        
        return jsonify({'status': 'success', 'message': 'Lançamento desconsiderado com sucesso.'})
        
    return jsonify({'status': 'error', 'message': 'Erro ao processar o cancelamento.'}), 400


@app.route('/request_token', methods=['POST'])
@login_required
def request_token():
    data = request.json
    valor_str = data.get('valor', '0')
    sem_movimento = data.get('sem_movimento', False)
    data_retirada = data.get('data')
    company_id = data.get('company_id')
    
    partner_name = data.get('partner_name')
    partner_cpf = data.get('partner_cpf')

    confirmed_tax = data.get('confirmed_tax', False)
    action = data.get('action', 'finish')

    if sem_movimento:
        valor_atual = 0.0
    else:
        try:
            valor_clean = valor_str.replace('R$', '').replace('.', '').replace(',', '.').strip()
            valor_atual = float(valor_clean)
        except:
            return jsonify({'status': 'error', 'message': 'Valor inválido.'}), 400

    if ((not sem_movimento and not valor_atual) or not data_retirada or not company_id):
        return jsonify({'status': 'error', 'message': 'Preencha todos os campos.'}), 400
        
    if not partner_name or not partner_cpf:
        return jsonify({'status': 'error', 'message': 'Sócio não selecionado.'}), 400

    company = mongo.db.companies.find_one({"_id": ObjectId(company_id), "user_id": ObjectId(current_user.id)})
    if not company:
        return jsonify({'status': 'error', 'message': 'Empresa inválida.'}), 400

    # Limpa o CPF para comparação blindada CNPJ + CPF
    partner_cpf_clean = re.sub(r'[^0-9]', '', partner_cpf)

    dt_obj = datetime.strptime(data_retirada, '%Y-%m-%d')
    start_date = datetime(dt_obj.year, dt_obj.month, 1)
    if dt_obj.month == 12:
        end_date = datetime(dt_obj.year + 1, 1, 1)
    else:
        end_date = datetime(dt_obj.year, dt_obj.month + 1, 1)

    pipeline = [
        {
            "$match": {
                "user_id": ObjectId(current_user.id),
                "company_cnpj": company['cnpj'],
                "data_retirada": {
                    "$gte": start_date.strftime('%Y-%m-%d'),
                    "$lt": end_date.strftime('%Y-%m-%d')
                },
                "status": {"$ne": "desconsiderado"}
            }
        }
    ]
    historico_records = list(mongo.db.user_financials.aggregate(pipeline))
    
    # Soma histórico apenas deste CPF E DESTA Empresa
    historico_total = 0.0
    for rec in historico_records:
        rec_cpf = re.sub(r'[^0-9]', '', rec.get('socio_cpf') or current_user.cpf)
        if rec_cpf == partner_cpf_clean:
            historico_total += rec.get('valor_numerico', 0)

    # Soma itens na sessão (lote temporário)
    batch_items = session.get('batch_items', [])
    batch_total = 0.0
    for item in batch_items:
        match_company = item.get('company_id') == company_id or item.get('company_cnpj') == company['cnpj']
        item_cpf_clean = re.sub(r'[^0-9]', '', item.get('partner_cpf', ''))
        if match_company and item_cpf_clean == partner_cpf_clean:
            batch_total += item.get('valor_numerico', 0)

    total_projetado = historico_total + batch_total + valor_atual
    tax_details = None

    if total_projetado > 50000:
        if not confirmed_tax:
            base_calculo = total_projetado / 0.9
            imposto = base_calculo * 0.10
            liquido = total_projetado

            return jsonify({
                'status': 'warning_tax',
                'message': 'Limite de isenção excedido.',
                'calculations': {
                    'total_acumulado': f"R$ {total_projetado:,.2f}",
                    'base_calculo': f"R$ {base_calculo:,.2f}",
                    'imposto': f"R$ {imposto:,.2f}",
                    'liquido': f"R$ {liquido:,.2f}",
                    'aviso': 'O valor total de retiradas deste sócio no mês para esta empresa excede R$ 50.000,00.'
                }
            })
        else:
            tax_details = {
                "base_calculo": total_projetado / 0.9,
                "imposto": (total_projetado / 0.9) * 0.10,
                "liquido_final": total_projetado,
                "total_acumulado_mes": total_projetado
            }

    current_item = {
        'valor_formatado': valor_str,
        'valor_numerico': valor_atual,
        'data_retirada': data_retirada,
        'company_id': str(company['_id']),
        'company_name': company['name'],
        'company_cnpj': company['cnpj'],
        'partner_name': partner_name, 
        'partner_cpf': partner_cpf_clean, 
        'timestamp': datetime.utcnow().isoformat()
    }

    if action == 'add':
        batch_items.append(current_item)
        session['batch_items'] = batch_items
        return jsonify({'status': 'added', 'message': 'Empresa adicionada à lista. Insira a próxima.'})

    elif action == 'finish':
        batch_items.append(current_item)

        session['final_submission'] = {
            'items': batch_items,
            'tax_details': tax_details
        }

        token = generate_token()
        session['auth_token'] = token

        try:
            content = f"""
            <div class="highlight-box">
                <span style="font-size:12px; font-weight:bold; color:#999; text-transform:uppercase;">Código de Assinatura</span>
                <span class="token-code">{token}</span>
            </div>
            <p style="text-align:center;">Este código valida <strong>{len(batch_items)}</strong> declaração(ões) pendente(s).</p>
            """
            msg = Message("Token Scryta", recipients=[current_user.email])
            msg.html = get_email_template("Token de segurança", "Assinatura em Lote", content)
            send_email_with_logo(msg)
            return jsonify({'status': 'token_sent'})
        except Exception as e:
            return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/submit_withdrawal', methods=['POST'])
@login_required
def submit_withdrawal():
    data = request.json
    code = data.get('code')

    if code == session.get('auth_token'):
        sub_data = session.get('final_submission')
        if not sub_data: return jsonify({'status': 'error', 'message': 'Sessão expirada.'}), 400

        items = sub_data['items']
        tax_details = sub_data.get('tax_details')

        batch_id_db = generate_token()

        rows_html = ""

        for item in items:
            s_nome = item.get('partner_name') or current_user.name
            s_cpf = re.sub(r'[^0-9]', '', (item.get('partner_cpf') or current_user.cpf))

            doc = {
                "user_id": ObjectId(current_user.id),
                "user_name": current_user.name, 
                "user_cpf": current_user.cpf,
                "user_email": current_user.email,
                "company_name": item['company_name'],
                "company_cnpj": item['company_cnpj'],
                "valor": item['valor_formatado'],
                "valor_numerico": item['valor_numerico'],
                "data_retirada": item['data_retirada'],
                "submitted_at": datetime.utcnow(),
                "ip_address": request.remote_addr,
                "validation_token_used": True,
                "batch_id": batch_id_db,
                "fiscal_data_ref": tax_details,
                "socio_nome": s_nome, 
                "socio_cpf": s_cpf,   
                "status": "ativo",
                "visualizado": False,
                "alerta_cancelamento": False
            }
            mongo.db.user_financials.insert_one(doc)

            date_ptbr = datetime.strptime(item['data_retirada'], '%Y-%m-%d').strftime('%d/%m/%Y')
            rows_html += f"""
            <tr class="receipt-row">
                <td style="text-align:left;">{item['company_cnpj']}<br><span style="font-size:11px;color:#999;">Sócio: {s_nome}</span></td>
                <td>{date_ptbr}</td>
                <td class="receipt-value">{item['valor_formatado']}</td>
            </tr>
            """

        try:
            extra_html = ""
            if tax_details:
                extra_html = f"""
                <br>
                <div style="background-color:#fffbf0; border:1px solid #ffeeba; border-radius:8px; padding:15px; margin-top:20px;">
                    <h3 style="margin:0 0 10px 0; font-size:14px; text-transform:uppercase; color:#FBBA00;">Cálculo Fiscal Consolidado</h3>
                    <table style="width:100%; font-size:13px;">
                        <tr><td style="padding:5px 0;">Total Acumulado</td><td style="text-align:right; font-weight:bold;">R$ {tax_details['total_acumulado_mes']:,.2f}</td></tr>
                        <tr><td style="padding:5px 0;">Base (Gross-up)</td><td style="text-align:right; font-weight:bold;">R$ {tax_details['base_calculo']:,.2f}</td></tr>
                        <tr><td style="padding:5px 0;">IRRF (10%)</td><td style="text-align:right; font-weight:bold; color:red;">- R$ {tax_details['imposto']:,.2f}</td></tr>
                        <tr style="border-top:1px solid #ddd;"><td style="padding:8px 0; font-weight:bold;">Líquido Final</td><td style="text-align:right; font-weight:bold; color:green;">R$ {tax_details['liquido_final']:,.2f}</td></tr>
                    </table>
                </div>
                """

            receipt_html = f"""
            <p>Confirmamos o processamento de <strong>{len(items)}</strong> declaração(ões).</p>
            <table class="receipt-table">
                <thead><tr><th style="text-align:left; padding:10px;">Empresa</th><th>Data</th><th style="text-align:right;">Valor</th></tr></thead>
                <tbody>{rows_html}</tbody>
            </table>
            {extra_html}
            <p style="font-size:11px; color:#999; margin-top:20px; text-align:center;">ID do Lote: {batch_id_db}</p>
            """
            msg = Message("Comprovante - Scryta", recipients=[current_user.email])
            msg.html = get_email_template("Envio confirmado.", "Recibo Oficial", receipt_html)
            send_email_with_logo(msg)
        except Exception as e:
            print(f"Erro email: {e}")

        session.pop('auth_token', None)
        session.pop('final_submission', None)
        session.pop('batch_items', None)

        return jsonify({'status': 'success'})

    return jsonify({'status': 'error', 'message': 'Token inválido.'}), 400


@app.route('/admin')
@login_required
def admin_panel():
    if not current_user.is_admin: return redirect(url_for('dashboard'))

    tab = request.args.get('tab', 'envios')

    if tab == 'envios':
        pipeline = [
            {
                "$addFields": {
                    "mes_ref": {"$substr": ["$data_retirada", 0, 7]}
                }
            },
            {
                "$sort": {"submitted_at": -1}
            },
            {
                "$lookup": {
                    "from": "companies",
                    "localField": "company_cnpj",
                    "foreignField": "cnpj",
                    "as": "company_info"
                }
            },
            {
                "$addFields": {
                    "real_company_name": {
                        "$cond": {
                            "if": {"$gt": [{"$size": "$company_info"}, 0]},
                            "then": {"$arrayElemAt": ["$company_info.name", 0]},
                            "else": "$company_name"
                        }
                    }
                }
            },
            {
                "$group": {
                    "_id": {
                        "company_cnpj": "$company_cnpj",
                        "mes_ref": "$mes_ref"
                    },
                    "company_name": {"$first": "$real_company_name"},
                    "user_name": {"$first": "$user_name"},
                    "user_cpf": {"$first": "$user_cpf"},
                    
                    "total_declarado": {
                        "$sum": {
                            "$cond": [
                                {"$ne": ["$status", "desconsiderado"]}, 
                                "$valor_numerico", 
                                0
                            ]
                        }
                    },
                    "qtd_empresas": {
                        "$sum": {
                            "$cond": [
                                {"$ne": ["$status", "desconsiderado"]}, 
                                1, 
                                0
                            ]
                        }
                    },
                    
                    "detalhes": {
                        "$push": {
                            "id": {"$toString": "$_id"},
                            "empresa": "$real_company_name",
                            "cnpj": "$company_cnpj",
                            "valor": "$valor",
                            "valor_numerico": "$valor_numerico",
                            "data": "$data_retirada",
                            "submitted_at": "$submitted_at",
                            "socio_nome": {"$ifNull": ["$socio_nome", "$user_name"]},
                            "socio_cpf": {"$ifNull": ["$socio_cpf", "$user_cpf"]},
                            "status": {"$ifNull": ["$status", "ativo"]},
                            "visualizado": {"$ifNull": ["$visualizado", False]},
                            "alerta_cancelamento": {"$ifNull": ["$alerta_cancelamento", False]}
                        }
                    }
                }
            }
        ]

        grouped_submissions = list(mongo.db.user_financials.aggregate(pipeline))

        for item in grouped_submissions:
            has_alerta = False
            has_pendente = False
            pendentes_count = 0
            
            # CÁLCULO MESTRE: Soma por CPF DENTRO deste Lote/Empresa/Mês
            cpf_totals = {}
            for det in item['detalhes']:
                if det.get('status') != 'desconsiderado':
                    cpf = re.sub(r'[^0-9]', '', det.get('socio_cpf') or '')
                    val = det.get('valor_numerico', 0)
                    cpf_totals[cpf] = cpf_totals.get(cpf, 0) + val
                    
            imposto_lote = 0

            for det in item['detalhes']:
                d_ret = det.get('data', '')
                det['retirada_fmt'] = f"{d_ret[8:10]}/{d_ret[5:7]}/{d_ret[2:4]}" if len(d_ret) == 10 else d_ret
                
                if det.get('submitted_at'):
                    det['envio_fmt'] = det['submitted_at'].strftime('%d/%m/%y')
                else:
                    det['envio_fmt'] = 'N/A'
            
            for det in item['detalhes']:
                if det.get('alerta_cancelamento'):
                    has_alerta = True
                if not det.get('visualizado') and det.get('status') != 'desconsiderado':
                    has_pendente = True
                    pendentes_count += 1
                
                if det.get('status') != 'desconsiderado':
                    cpf = re.sub(r'[^0-9]', '', det.get('socio_cpf') or '')
                    if cpf_totals.get(cpf, 0) > 50000:
                        val = det.get('valor_numerico', 0)
                        det['irrf'] = (val / 0.9) * 0.10
                        imposto_lote += det['irrf']
                    else:
                        det['irrf'] = 0
                else:
                    det['irrf'] = 0
                    
            if has_alerta:
                item['row_status'] = 'ALERTA'
                item['sort_order'] = 1
            elif has_pendente:
                item['row_status'] = 'PENDENTE'
                item['cont_pendentes'] = pendentes_count
                item['sort_order'] = 2
            else:
                item['row_status'] = 'VALIDADO'
                item['sort_order'] = 3

            if imposto_lote > 0:
                item['calculo_dinamico'] = {
                    "imposto": imposto_lote
                }
            else:
                item['calculo_dinamico'] = None

        grouped_submissions.sort(key=lambda x: x['_id']['mes_ref'], reverse=True)
        grouped_submissions.sort(key=lambda x: x['sort_order'])

        return render_template('admin.html', grouped_data=grouped_submissions, active_tab='envios')

    else:
        users_records = list(mongo.db.users.find().sort("created_at", -1))
        return render_template('admin.html', users=users_records, active_tab='users')


@app.route('/admin/toggle_user/<user_id>', methods=['POST'])
@login_required
def admin_toggle_user(user_id):
    if not current_user.is_admin: 
        return redirect(url_for('dashboard'))

    user = mongo.db.users.find_one({"_id": ObjectId(user_id)})
    if user:
        new_status = not user.get('is_admin', False)
        mongo.db.users.update_one({"_id": ObjectId(user_id)}, {"$set": {"is_admin": new_status}})
        flash('Privilégios de administrador atualizados com sucesso.', 'success')
        
    return redirect(url_for('admin_panel', tab='users'))


@app.route('/admin/record/toggle_view/<record_id>', methods=['POST'])
@login_required
def toggle_record_view(record_id):
    if not current_user.is_admin:
        return jsonify({'status': 'error', 'message': 'Não autorizado'}), 403

    record = mongo.db.user_financials.find_one({"_id": ObjectId(record_id)})
    if not record:
        return jsonify({'status': 'error', 'message': 'Registro não encontrado'}), 404

    current_visualizado = record.get('visualizado', False)
    new_visualizado = not current_visualizado
    is_desconsiderado = record.get('status') == 'desconsiderado'

    mongo.db.user_financials.update_one(
        {"_id": ObjectId(record_id)},
        {"$set": {
            "visualizado": new_visualizado,
            "alerta_cancelamento": False
        }}
    )

    return jsonify({
        'status': 'success',
        'visualizado': new_visualizado,
        'is_desconsiderado': is_desconsiderado
    })


@app.route('/admin/term_proof/<user_id>')
@login_required
def term_proof(user_id):
    if not current_user.is_admin: return redirect(url_for('dashboard'))

    user = mongo.db.users.find_one({"_id": ObjectId(user_id)})
    if not user: return "Usuário não encontrado", 404

    financials = list(mongo.db.user_financials.find({"user_id": ObjectId(user_id)}).sort("data_retirada", -1))

    return render_template('print_proof.html', user=user, financials=financials, now=datetime.utcnow())


@app.route('/admin/export')
@login_required
def export_excel():
    if not current_user.is_admin: return redirect(url_for('dashboard'))

    wb = openpyxl.Workbook()

    ws_resumo = wb.active
    ws_resumo.title = "Total Lotes"
    ws_detalhes = wb.create_sheet(title="Registros Individuais")

    header_fill = PatternFill(start_color="413D3A", end_color="413D3A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border_style = Border(
        left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB')
    )
    center_align = Alignment(horizontal="center", vertical="center")

    headers_resumo = ['Mês Referência', 'Nome da Empresa', 'CNPJ', 'Total Acumulado Lote', 'Base Cálculo', 'Imposto Retido', 'Líquido Recebido']
    ws_resumo.append(headers_resumo)
    for cell in ws_resumo[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Aba de Resumo
    pipeline = [
        {"$addFields": {"mes_ref": {"$substr": ["$data_retirada", 0, 7]}}},
        {
            "$lookup": {
                "from": "companies",
                "localField": "company_cnpj",
                "foreignField": "cnpj",
                "as": "company_info"
            }
        },
        {
            "$addFields": {
                "real_company_name": {
                    "$cond": {
                        "if": {"$gt": [{"$size": "$company_info"}, 0]},
                        "then": {"$arrayElemAt": ["$company_info.name", 0]},
                        "else": "$company_name"
                    }
                }
            }
        },
        {"$group": {
            "_id": {"company_cnpj": "$company_cnpj", "mes_ref": "$mes_ref"},
            "company_name": {"$first": "$real_company_name"},
            "user_name": {"$first": "$user_name"},
            "user_cpf": {"$first": "$user_cpf"},
            "detalhes": {
                "$push": {
                    "valor_numerico": "$valor_numerico",
                    "socio_cpf": {"$ifNull": ["$socio_cpf", "$user_cpf"]},
                    "status": {"$ifNull": ["$status", "ativo"]}
                }
            }
        }},
        {"$sort": {"_id.mes_ref": -1}}
    ]
    resumo_data = list(mongo.db.user_financials.aggregate(pipeline))

    for r in resumo_data:
        cpf_totals = {}
        total_lote = 0
        for det in r['detalhes']:
            if det.get('status') != 'desconsiderado':
                cpf = re.sub(r'[^0-9]', '', det.get('socio_cpf') or '')
                val = det.get('valor_numerico', 0)
                cpf_totals[cpf] = cpf_totals.get(cpf, 0) + val
                total_lote += val
        
        imposto = 0
        base = 0
        for det in r['detalhes']:
            if det.get('status') != 'desconsiderado':
                cpf = re.sub(r'[^0-9]', '', det.get('socio_cpf') or '')
                if cpf_totals.get(cpf, 0) > 50000:
                    val = det.get('valor_numerico', 0)
                    imposto += (val / 0.9) * 0.10
                    base += (val / 0.9)
                    
        company_cnpj = r['_id'].get('company_cnpj')
        company_name = r.get('company_name', '')
        row = [r['_id']['mes_ref'], company_name or '', company_cnpj or '', total_lote, base, imposto, total_lote]
        ws_resumo.append(row)

    for row in ws_resumo.iter_rows(min_row=2, max_col=7):
        for cell in row: cell.border = border_style
        for idx in [3, 4, 5, 6]: row[idx].number_format = '"R$" #,##0.00'
        
    for col in ws_resumo.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_resumo.column_dimensions[column].width = max_length + 6

    # Aba de Registros Individuais
    headers_detalhes = ['ID Lote', 'CNPJ Empresa', 'Nome Sócio/Titular', 'CPF Sócio/Titular', 'Data Lançamento', 'Valor Lançamento', 'IRRF Retido']
    ws_detalhes.append(headers_detalhes)
    for cell in ws_detalhes[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Pré-calcula os totais (Mês + CNPJ + CPF) para aplicar nas linhas individuais
    all_active = list(mongo.db.user_financials.find({"status": {"$ne": "desconsiderado"}}))
    totals_dict = {}
    for rec in all_active:
        mes_ref = rec.get('data_retirada', '')[:7]
        cnpj = rec.get('company_cnpj', '')
        cpf = re.sub(r'[^0-9]', '', rec.get('socio_cpf') or rec.get('user_cpf') or '')
        val = rec.get('valor_numerico', 0)
        key = (mes_ref, cnpj, cpf)
        totals_dict[key] = totals_dict.get(key, 0) + val

    financial_records = list(mongo.db.user_financials.find().sort([("data_retirada", -1), ("company_cnpj", 1)]))
    for s in financial_records:
        mes_ref = s.get('data_retirada', '')[:7]
        cnpj = s.get('company_cnpj', '')
        cpf = re.sub(r'[^0-9]', '', s.get('socio_cpf') or s.get('user_cpf') or '')
        val = s.get('valor_numerico', 0)
        status = s.get('status', 'ativo')

        imposto_individual = 0
        if status != 'desconsiderado':
            if totals_dict.get((mes_ref, cnpj, cpf), 0) > 50000:
                imposto_individual = (val / 0.9) * 0.10

        row = [
            s.get('batch_id', 'N/A'),
            cnpj,
            s.get('socio_nome', s.get('user_name', '')), 
            cpf,
            s.get('data_retirada', ''),
            val,
            imposto_individual
        ]
        ws_detalhes.append(row)

    for row in ws_detalhes.iter_rows(min_row=2, max_col=7):
        for cell in row: cell.border = border_style
        row[5].number_format = '"R$" #,##0.00'
        row[6].number_format = '"R$" #,##0.00'

    for col in ws_detalhes.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_detalhes.column_dimensions[column].width = max_length + 6

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = make_response(output.getvalue())
    resp.headers["Content-Disposition"] = "attachment; filename=relatorio_completo_scryta.xlsx"
    resp.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return resp


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5894)